"""Century-scale market data, so a conclusion can be tested outside the
one window our whole sample lives in.

Everything else in this repository runs on 2005-2026. That window
contains exactly one great financial crisis, one rate cycle worth the
name, and one seventeen-year bull market with a pandemic in the middle
of it. Every claim we have made about drawdowns, about trend rules,
about what a defensive sleeve is for, has an n of 1 hanging off it. The
files fetched here are how that gets tested: Shiller's monthly series
reaches 1871 and Damodaran's annual returns reach 1928, which between
them cover 1929, 1937, 1973-74, 1987 and the whole of the inflation of
the seventies — five regimes our sample has never seen.

**What these datasets can and cannot do, stated before anything else,
because the temptation runs entirely one way.** They are INDEX-LEVEL.
There is no cross-section anywhere in this file: no constituent list,
no per-security row, no delisting date, no ticker. You can ask whether
a market-timing rule survives 1929 and 1974. You cannot test stock
selection with them at all — not badly, not approximately, not with
care. There is nothing to select from. A backtest that reads these and
reports a per-name result has a bug.

They are also SLOW. Shiller is monthly, Damodaran is annual, and the
FRED series here are monthly. A daily rule cannot be evaluated on them;
the most you get is the monthly version of the rule, which is a
different rule.

**Three traps, each of which produces a clean-looking frame.**

*The Shiller date column is not a decimal year.* It is encoded YYYY.MM,
so 1871.01 is January and 1871.10 is OCTOBER — and Excel stores that
last one as the float 1871.1, which every naive parse reads as January.
The collapse is silent: you get two Januaries and no October, a 1,867
row file becomes 1,867 rows still, and the series just quietly runs a
month wrong for a quarter of every year. `shiller_month` decodes it and
cross-checks against the workbook's own Date Fraction column, and the
parse then asserts the months form an unbroken monthly sequence — which
is the check that actually fires, because a merged October shows up as
a duplicated January and a hole where October should be.

*The Yale URL everybody cites is frozen.* `econ.yale.edu/~shiller/data/
ie_data.xls` still answers 200 with a well-formed workbook whose last
row is September 2023. The live file moved to shillerdata.com. Nothing
in the stale copy says it is stale, so a sample built from it is simply
three years short and looks complete. It is exported here as
`SHILLER_YALE_MIRROR` with that end date named, and it is deliberately
NOT an automatic fallback: silently substituting a truncated history
for an unreachable current one is the exact failure this repository
exists to prevent.

*Damodaran's monthly ERP sheet is hand-maintained and about one cell in
a hundred is a string.* Not a note — a number typed as text: `'3.90%'`,
`'84,88'` with a European decimal comma, `'175.51*'` with a footnote
marker, and one whole month (September 2024) where every field is
formatted that way. `pd.to_numeric(errors="coerce")` NaNs that month
without a word. `coerce_number` repairs what is repairable, returns NaN
for a cell that is genuinely prose (`'Ended'`), and the parser warns
with a count so a jump in that count is visible.

**The raw bytes are what gets cached, not the parse.** These are
individually maintained academic workbooks: they move, they change
column layout between years, and they are occasionally down. The parse
is the part we expect to have to change; the download is the part we
must not repeat. So the cache holds the workbook itself, keyed on its
URL, and a re-parse costs nothing. It also never expires on a timer — a
file its author updates once each January, re-downloaded nightly, is
rudeness that returns no information. `refresh=True` is how you ask for
a new vintage, and it is meant to be a decision somebody made.

That sha256 in the cache sidecar matters more than it looks.
Damodaran's own usage page says that if he decides he has been
computing something incorrectly for ten years he has "no qualms about
changing the way I do it" — so the 1928 return in this January's file
need not equal the 1928 return in the last one. These series are a
CURRENT-VINTAGE reconstruction of history, not a point-in-time record of
what anyone could have known. The digest is the vintage id, and the
cached bytes are what makes a result from them reproducible at all.

There is deliberately no `DataSource` subclass here. That contract is
about a securities panel — a security master, delistings, as-reported
fundamentals — and every one of its methods would have to raise or lie.
An index level implementing an interface built for a cross-section is a
frame that will eventually be joined to one.
"""

from __future__ import annotations

import hashlib
import io
import math
import re
import time
import warnings
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd
import requests

from .base import SourceUnavailable
from .tbill import TbillUnavailable, fetch_observations

if TYPE_CHECKING:
    from .cache import ParquetCache


#: Carries an address, because these are somebody's personal academic
#: web servers rather than an API with a support desk. A host that
#: wants us to stop should be able to work out who to tell.
USER_AGENT = (
    "GriffinFund-Research/0.1 (academic backtest; newtheyork@gmail.com)"
)

#: The live Shiller workbook. Long, opaque, and hosted on a website
#: builder's CDN — which is exactly why it is pinned here rather than
#: rediscovered by scraping shillerdata.com on every run: a fetch that
#: depends on a page's markup breaks the day the page is restyled, and
#: it breaks by finding nothing rather than by failing.
SHILLER_URL = (
    "https://img1.wsimg.com/blobby/go/e5e77e0b-59d1-44d9-ab25-4763ac982e53"
    "/downloads/165d8a6e-26bf-44ec-a26c-a35f7f993480/ie_data.xls"
)

#: The URL in every paper and every Stack Overflow answer. It still
#: serves a valid, parseable, ENTIRELY STALE workbook ending September
#: 2023. Exported so a reader who has one lying around can identify it,
#: never wired up as a fallback — see the module docstring.
SHILLER_YALE_MIRROR = "http://www.econ.yale.edu/~shiller/data/ie_data.xls"
SHILLER_YALE_MIRROR_LAST_ROW = "2023-09"

DAMODARAN_RETURNS_URL = (
    "https://pages.stern.nyu.edu/~adamodar/pc/datasets/histretSP.xls"
)
DAMODARAN_IMPLIED_ERP_URL = (
    "https://pages.stern.nyu.edu/~adamodar/pc/datasets/histimpl.xls"
)
DAMODARAN_ERP_MONTHLY_URL = (
    "https://pages.stern.nyu.edu/~adamodar/pc/implprem/ERPbymonth.xlsx"
)

#: Slower than the half-second we allow a paid vendor's API. One
#: workbook a run is not a load on anybody, and the cost of being
#: remembered as impolite by a server with no rate-limit documentation
#: is losing the source entirely.
MIN_REQUEST_INTERVAL_SECONDS = 2.0
MAX_ATTEMPTS = 3
BACKOFF_BASE_SECONDS = 3.0
BACKOFF_CAP_SECONDS = 30.0
RETRY_STATUSES = frozenset({429, *range(500, 512)})

#: First bytes of an OLE2 compound document (legacy .xls) and of a zip
#: (.xlsx). Checked before the file reaches a parser, because the
#: characteristic failure of a personal web host is HTTP 200 carrying an
#: error page, and "xlrd cannot open this document" sends somebody to
#: debug the wrong layer.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP_MAGIC = b"PK\x03\x04"


class LongHistoryUnavailable(SourceUnavailable):
    """A workbook could not be fetched, opened, or understood.

    Deliberately never an empty frame. These files are the only reason
    we can say anything about 1929 at all, and a run that quietly
    produced no rows for them would report the modern sample's
    conclusions as though they had been tested against a century.
    """


# -- what each dataset is, and what it is not -----------------------------


@dataclass(frozen=True)
class LongDataset:
    """One published series, described honestly enough to cite.

    Every field here is a fact about the SOURCE rather than about the
    frame, and the two that carry the weight are `survivorship` and
    `cannot`. A reader who takes only the column names away from this
    module will assume an index level behaves like a portfolio they
    could have held. Several of these do not.
    """

    key: str
    name: str
    author: str
    url: str
    frequency: str
    #: The first observation actually present, verified by reading the
    #: file rather than by repeating the author's prose.
    starts: str
    #: The last observation in the vintage read on 2026-08-02. Recorded
    #: so a stale mirror can be recognised on sight.
    observed_through: str
    cadence: str
    survivorship: str
    redistribution: str
    can: tuple[str, ...]
    cannot: tuple[str, ...]


_SHILLER_SURVIVORSHIP = (
    "Survivorship-free in the only sense an index level can be, and "
    "unable to answer the question in any other sense. A company that "
    "failed took the index down with it in the month it failed and was "
    "then replaced, so the level series contains both events and no "
    "dead name has been dropped out of it — that is a portfolio-level "
    "guarantee and it is real. What it cannot do is let you check: "
    "there is no constituent list, no delisting date, nothing to look "
    "up. Two further caveats which are selection rather than "
    "survivorship, and which we cannot test from here: prices before "
    "1926 are the Cowles Commission's retrospective reconstruction "
    "(Common Stock Indexes, 1939), compiled from the records that "
    "survived to 1939 rather than published at the time; and dividends "
    "and earnings before 1926 are interpolated from Cowles annual "
    "figures, so their monthly variation is arithmetic, not history."
)

_DAMODARAN_SURVIVORSHIP = (
    "Index and portfolio level, so survivorship-free in the same "
    "limited sense and unverifiable in the same way. Two components "
    "deserve their own sentence. The small-cap decile descends from "
    "CRSP by way of Ken French, and CRSP retains securities that "
    "stopped trading, so that leg is survivorship-free by "
    "construction rather than by assumption. The 10-year Treasury "
    "return is NOT an observed total-return index: Damodaran computes "
    "it from the constant-maturity yield with his own coupon and price "
    "arithmetic, so it is a construction, and a bond backtest run on "
    "it is testing his construction as much as the market."
)

_VINTAGE_WARNING = (
    "Not point-in-time. The author revises method and inputs between "
    "annual vintages — his own usage page says he has no qualms about "
    "changing how he computes something he has been computing for ten "
    "years — so a figure for 1928 in this vintage need not match the "
    "same figure in a file downloaded three years ago. The cached "
    "bytes and their sha256 are the vintage id."
)

DATASETS: Mapping[str, LongDataset] = {
    "shiller": LongDataset(
        key="shiller",
        name="US stock prices, dividends, earnings, CPI, long rate, CAPE",
        author="Robert J. Shiller (shillerdata.com)",
        url=SHILLER_URL,
        frequency="monthly",
        starts="1871-01",
        observed_through="2026-07",
        cadence=(
            "Irregular, roughly monthly. The author posts a new "
            "workbook when he updates it; there is no schedule and no "
            "changelog."
        ),
        survivorship=_SHILLER_SURVIVORSHIP,
        redistribution=(
            "No licence is published. The site offers the file for "
            "download and the workbook carries a disclaimer of "
            "accuracy and completeness; nothing grants or refuses "
            "redistribution. Treated here as free to read and cite, "
            "and not ours to republish — quant/data/ is gitignored, so "
            "no copy of it enters this repository."
        ),
        can=(
            "Test a monthly market-timing or valuation rule across "
            "1929, 1937, 1973-74, 1987 and the 1970s inflation.",
            "Supply a real (CPI-deflated) equity total return from "
            "1871, and a CAPE from 1881.",
            "Supply a long rate (GS10) monthly from 1871, which is "
            "eighty years earlier than FRED's own GS10 starts.",
        ),
        cannot=(
            "Say anything about individual securities. There is no "
            "cross-section in this file at all.",
            "Support a daily or weekly rule.",
            "Be used as a month-end price series. See the note on "
            "monthly averaging in `parse_shiller` — the price column "
            "is an average of the month's daily closes, and using it "
            "as a close manufactures the serial correlation a trend "
            "rule is looking for.",
        ),
    ),
    "damodaran_returns": LongDataset(
        key="damodaran_returns",
        name="Annual returns on US stocks, bonds, bills, Baa, property, gold",
        author="Aswath Damodaran (NYU Stern), histretSP.xls",
        url=DAMODARAN_RETURNS_URL,
        frequency="annual",
        starts="1928",
        observed_through="2025",
        cadence=(
            "Once a year, in the first two weeks of January. The "
            "vintage read here was published 2026-01-09."
        ),
        survivorship=_DAMODARAN_SURVIVORSHIP + " " + _VINTAGE_WARNING,
        redistribution=(
            'The author states: "I hope you find this data useful and '
            'there are no strings attached", and under Usage Rules, '
            '"I am not good at making rules and thus have very few '
            'related to the use of my data... If you do use my data '
            'and wish to acknowledge that you did get the data off my '
            'site, I thank you." Free for academic use; attribution '
            "requested rather than required."
        ),
        can=(
            "Give a 98-year record of the equity risk premium against "
            "both bills and bonds, arithmetic and geometric.",
            "Put a single bad year in a century of context, in real "
            "terms as well as nominal.",
            "Serve as an independent second witness to Shiller over "
            "1928-2025 — see `reconcile_annual_equity_return`.",
        ),
        cannot=(
            "Resolve anything finer than a calendar year.",
            "Say anything about individual securities.",
            "Be treated as what an investor saw at the time. It is "
            "this January's reconstruction of history.",
        ),
    ),
    "damodaran_implied_erp": LongDataset(
        key="damodaran_implied_erp",
        name="Implied equity risk premium, annual",
        author="Aswath Damodaran (NYU Stern), histimpl.xls",
        url=DAMODARAN_IMPLIED_ERP_URL,
        frequency="annual",
        starts="1960",
        observed_through="2025",
        cadence="Once a year, alongside the returns file.",
        survivorship=(
            "Not applicable — this is a forward-looking discount rate "
            "backed out of the index level, its cash yield and a "
            "growth assumption, not a return anyone earned. "
            + _VINTAGE_WARNING
        ),
        redistribution="As above: no strings attached, attribution welcomed.",
        can=(
            "Show what the market was pricing as compensation for "
            "equity risk at each year end since 1960, which is a "
            "different and more honest question than what equities "
            "went on to return.",
        ),
        cannot=(
            "Reach back a century — it starts in 1960, so it does not "
            "cover 1929 or 1974 and cannot help with either.",
            "Be read as a forecast. It is the premium that would "
            "justify the index at that moment given the assumptions "
            "in the model, and the assumptions are the author's.",
        ),
    ),
    "damodaran_erp_monthly": LongDataset(
        key="damodaran_erp_monthly",
        name="Implied equity risk premium, monthly",
        author="Aswath Damodaran (NYU Stern), ERPbymonth.xlsx",
        url=DAMODARAN_ERP_MONTHLY_URL,
        frequency="monthly",
        starts="2008-09",
        observed_through="2026-08",
        cadence="Monthly, at the start of each month.",
        survivorship="Not applicable — a discount rate, not a return.",
        redistribution="As above: no strings attached, attribution welcomed.",
        can=(
            "Track the priced equity premium month by month through "
            "the crisis of 2008-09 and everything since.",
        ),
        cannot=(
            "Be described as long-horizon data. It starts in September "
            "2008 and is the shortest series in this module — it is "
            "here because the brief asked for the implied ERP, not "
            "because eighteen years tests anything.",
            "Be read cell by cell without the coercion in "
            "`coerce_number`. About one value in a hundred is a "
            "number typed as text.",
        ),
    ),
    "fred_macro": LongDataset(
        key="fred_macro",
        name="Long US macro series (recession dating, CPI, output, credit)",
        author="Federal Reserve Bank of St. Louis (FRED)",
        url="https://fred.stlouisfed.org/graph/fredgraph.csv",
        frequency="monthly",
        starts="1854-12 (USREC); each series differs, see FRED_CENTURY_SERIES",
        observed_through="2026-06",
        cadence="Each series on its own publication schedule.",
        survivorship=(
            "Not applicable — aggregate statistics, not a panel of "
            "securities. The relevant hazard is the opposite one and "
            "it is severe: USREC is NBER's dating, assigned "
            "RETROSPECTIVELY. The committee announces a peak a year or "
            "more after it happened, so the value for month t was not "
            "known in month t. USREC describes a regime after the "
            "fact; it may never be an input to a rule that trades. "
            "CPI and INDPRO are also revised after first publication, "
            "and what this endpoint serves is the current revision."
        ),
        redistribution=(
            "FRED serves these without a key. Most series here are US "
            "federal statistics and carry no copyright; FRED's own "
            "terms note that individual series may be copyrighted by "
            "their source, so check the series page before "
            "republishing one."
        ),
        can=(
            "Mark every US recession since 1854 on a chart.",
            "Supply CPI from 1913, industrial production and Moody's "
            "Aaa/Baa yields from 1919 — a credit spread through 1929 "
            "and 1932.",
        ),
        cannot=(
            "Be used as a real-time signal where the series is dated "
            "retrospectively (USREC) or revised (CPI, INDPRO).",
            "Say anything about individual securities.",
        ),
    ),
}


#: Keyless FRED series that reach further back than our sample, with
#: the first observation each one actually returned when this was
#: written. The dates are measured, not quoted from a documentation
#: page — a start date that is wrong by a decade is invisible until
#: somebody wonders why a chart begins where it does.
#:
#: `fredseries.py` is the general FRED reader and the place for
#: anything that needs a publication lag or an as-of vintage. This list
#: is only the century-scale slice — the series that reach back past
#: our 2005 sample, which is the one thing this module is for — and it
#: deliberately makes no as-of claim at all. If you find yourself
#: wanting to know what a number looked like on a past date, you are in
#: the wrong file.
FRED_CENTURY_SERIES: Mapping[str, str] = {
    "USREC": "NBER recession indicator (1 in recession), 1854-12",
    "CPIAUCNS": "CPI-U, all items, not seasonally adjusted, 1913-01",
    "INDPRO": "Industrial production index, 1919-01",
    "AAA": "Moody's seasoned Aaa corporate bond yield, 1919-01",
    "BAA": "Moody's seasoned Baa corporate bond yield, 1919-01",
    "UNRATE": "Civilian unemployment rate, 1948-01",
    "GS10": "10-year Treasury constant maturity yield, 1953-04",
    "FEDFUNDS": "Effective federal funds rate, 1954-07",
    "M2SL": "M2 money stock, seasonally adjusted, 1959-01",
}


def catalogue() -> pd.DataFrame:
    """Every dataset in this module as a frame, for a report to print.

    One row per source with its start, its cadence, its terms and its
    survivorship position. The `cannot` column is joined into the same
    frame rather than kept somewhere more tasteful precisely so it
    cannot be dropped when somebody renders a summary table.
    """
    rows = [
        {
            "key": d.key,
            "name": d.name,
            "author": d.author,
            "frequency": d.frequency,
            "starts": d.starts,
            "observed_through": d.observed_through,
            "cadence": d.cadence,
            "survivorship": d.survivorship,
            "redistribution": d.redistribution,
            "can": " | ".join(d.can),
            "cannot": " | ".join(d.cannot),
            "url": d.url,
        }
        for d in DATASETS.values()
    ]
    return pd.DataFrame(rows)


# -- the date encoding, which is the trap ---------------------------------


def shiller_month(value: Any, fraction: Any = None) -> tuple[int, int]:
    """Decode Shiller's YYYY.MM date cell into `(year, month)`.

    The cell is a NUMBER whose fractional part is the month in
    hundredths: 1871.01 is January and 1871.10 is October. Excel stores
    the second as the float 1871.1, so every string-shaped parse — a
    split on the decimal point, `%Y.%m`, taking the first digit after
    the dot — reads October as January. Multiplying the fraction by a
    hundred and rounding is what separates them, and it is the only
    step here that matters.

    `fraction` is the workbook's own Date Fraction column, which
    encodes the same month a completely different way: `year +
    (month - 0.5) / 12`. Where it is supplied the two are required to
    agree. That is not belt and braces — it is the only independent
    witness in the file, and a decoding this easy to get wrong should
    not be trusted against itself.
    """
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        raise ValueError(f"not a Shiller date cell: {value!r}")

    # The epsilon guards a year stored as 1871.9999999999 by a
    # spreadsheet that has done arithmetic on it; without it, floor
    # returns 1870 and every month is off by twelve.
    year = int(math.floor(number + 1e-9))
    month = int(round((number - year) * 100.0))
    if not 1 <= month <= 12:
        raise ValueError(
            f"Shiller date {number!r} decodes to month {month}, which is "
            f"not a month. The column is YYYY.MM, not a decimal year — a "
            f"value like 1871.5 would be a decimal year and is not what "
            f"this file publishes."
        )

    frac = _as_float(fraction)
    if frac is not None and math.isfinite(frac):
        from_fraction = int(round((frac - year) * 12.0 + 0.5))
        if from_fraction != month:
            raise ValueError(
                f"Shiller row {number!r} decodes to month {month} but its "
                f"Date Fraction {frac!r} says month {from_fraction}. The "
                f"two columns encode the same month independently, so a "
                f"disagreement means the file's layout changed and this "
                f"parser is reading the wrong column — not that one of "
                f"them is slightly off."
            )
    return year, month


# -- number coercion, for a sheet somebody maintains by hand --------------

_TRAILING_JUNK = re.compile(r"[^0-9]+$")
_LEADING_JUNK = re.compile(r"^[^0-9+\-.]+")
_DECIMAL_COMMA = re.compile(r"^[+-]?\d+,\d{1,2}$")


def coerce_number(value: Any) -> float:
    """A float from a cell that may have been typed as text.

    Damodaran's monthly ERP sheet is maintained by a person, and it
    shows: `'3.90%'`, `'84,88'`, `'175.51*'`, `'5648'`, and one month
    where every field is formatted that way. Each of those is a number
    the author meant; none of them survives `pd.to_numeric`, which
    turns the lot into NaN and takes a whole month of the series with
    it in silence.

    Four repairs, and the reasoning for the awkward one. A trailing `%`
    divides by a hundred, because the neighbouring rows in the same
    column are stored as 0.0372 rather than 3.72. A trailing footnote
    marker is dropped. A comma is a THOUSANDS separator except when it
    is the only comma, has one or two digits after it, and there is no
    decimal point in the string — then it is a European decimal comma,
    which is what `'84,88'` and `'6,36%'` are (the neighbouring values
    are 85 and 0.0628, so the reading is not a guess). `'1,234'` stays
    one thousand two hundred and thirty four under that rule, which is
    the right way round for an American author's spreadsheet.

    A parenthesised value is negative. No cell in either workbook uses
    the accounting convention today, and it is handled anyway because
    the failure if one ever does is a sign flip — a number that stays
    the right magnitude, passes every range check, and turns a loss
    into a gain.

    Anything genuinely not a number — `'Ended'`, a note, a blank —
    returns NaN. It has to: refusing the file over a word the author
    typed in a column he stopped maintaining would cost the series for
    no gain, and NaN is what "no value here" already means everywhere
    downstream.
    """
    if value is None:
        return float("nan")
    if isinstance(value, bool):
        # A bool is an int in Python and would arrive here as 1.0,
        # which is a plausible-looking rate. It is never what a
        # spreadsheet meant.
        return float("nan")
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    if not isinstance(value, str):
        return float("nan")

    text = value.strip()
    if not text:
        return float("nan")

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1].strip()

    percent = text.endswith("%")
    if percent:
        text = text[:-1].strip()

    text = _LEADING_JUNK.sub("", text)
    text = _TRAILING_JUNK.sub("", text)
    if not text:
        return float("nan")

    if _DECIMAL_COMMA.match(text):
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        number = float(text)
    except ValueError:
        return float("nan")
    if percent:
        number /= 100.0
    return -number if negative else number


def _as_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return None
    return None


# -- workbooks as grids ---------------------------------------------------
#
# Both parsers work on a plain list-of-rows-of-cells rather than on a
# workbook object. That is not indirection for its own sake: xlrd can
# only open .xls and openpyxl only .xlsx, so a parser written against
# either is untestable without shipping a binary fixture, and a binary
# fixture of somebody else's workbook goes stale without telling
# anyone. A grid is something a test can write down.

Grid = list[list[Any]]


def grid_from_xls(data: bytes, sheet: str) -> Grid:
    """One sheet of a legacy .xls (BIFF) workbook as a grid."""
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - environment
        raise LongHistoryUnavailable(
            "reading .xls needs xlrd, which is not installed. This is an "
            "unconfigured environment, not a source that went down: "
            "`uv pip install xlrd openpyxl`. Neither package is in "
            "pyproject.toml yet — add them there if this module becomes "
            "part of a run rather than an exploration."
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise LongHistoryUnavailable(
            f"could not open the workbook: {type(exc).__name__}: {exc}"
        ) from exc

    if sheet not in book.sheet_names():
        raise LongHistoryUnavailable(
            f"the workbook has no sheet named {sheet!r}; it has "
            f"{book.sheet_names()}. A renamed sheet is a layout change, "
            f"and guessing at the nearest one would read a different "
            f"series under the same column names."
        )
    view = book.sheet_by_name(sheet)
    return [
        [view.cell(r, c).value for c in range(view.ncols)]
        for r in range(view.nrows)
    ]


def grid_from_xlsx(data: bytes, sheet: str) -> Grid:
    """One sheet of an .xlsx workbook as a grid."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment
        raise LongHistoryUnavailable(
            "reading .xlsx needs openpyxl, which is not installed: "
            "`uv pip install xlrd openpyxl`."
        ) from exc

    try:
        with warnings.catch_warnings():
            # openpyxl warns about a spreadsheet extension it does not
            # implement and then reads the cells perfectly. Narrowly
            # silenced because it is about the author's file rather
            # than about our data, and it fires on every single load.
            warnings.simplefilter("ignore", UserWarning)
            book = openpyxl.load_workbook(
                io.BytesIO(data), data_only=True, read_only=True
            )
    except Exception as exc:
        raise LongHistoryUnavailable(
            f"could not open the workbook: {type(exc).__name__}: {exc}"
        ) from exc

    if sheet not in book.sheetnames:
        raise LongHistoryUnavailable(
            f"the workbook has no sheet named {sheet!r}; it has "
            f"{book.sheetnames}."
        )
    return [list(row) for row in book[sheet].iter_rows(values_only=True)]


# -- resolving columns by what is written above them ----------------------
#
# Both workbooks name a column with a STACK of text — Shiller runs a
# label down six rows, Damodaran writes a merged banner over each block
# — and in both files the bottom row alone is ambiguous. Damodaran's
# `S&P 500 (includes dividends)3` appears twice, once in the nominal
# $100 block and once in the real one, and only the banner tells them
# apart. Reading by position instead would survive today and put
# nominal wealth in the real column the first year he inserts a
# column.


@dataclass(frozen=True)
class _ColumnSpec:
    """A column named by what must and must not appear in its label."""

    name: str
    needs: tuple[str, ...]
    forbids: tuple[str, ...] = ()
    text: bool = False


def _norm(text: Any) -> str:
    """Lowercase, whitespace-collapsed, and `". "` closed up to `"."`.

    That last one is the only surprising rule and it earns its place:
    Damodaran writes `US T. Bond (10-year)` in one block and
    `!0-year T.Bonds` in another (the exclamation mark is his typo for
    a 1, and it is still there). Closing the space makes one needle,
    `t.bond`, match both.
    """
    if text is None:
        return ""
    out = " ".join(str(text).split()).lower()
    return out.replace(". ", ".")


def _column_labels(
    grid: Grid, rows: Sequence[int], *, forward_fill: bool = False
) -> list[str]:
    """The stacked header text per column, joined and normalised.

    `forward_fill` carries a merged banner rightward across the columns
    it spans, which is how Damodaran's block headers are stored — the
    text sits in the leftmost cell of the merge and every other cell in
    it is empty.
    """
    width = max((len(grid[r]) for r in rows if r < len(grid)), default=0)
    stacks: list[list[str]] = [[] for _ in range(width)]

    for r in rows:
        if r >= len(grid):
            continue
        row = grid[r]
        carried = ""
        for c in range(width):
            cell = row[c] if c < len(row) else None
            piece = " ".join(str(cell).split()) if isinstance(cell, str) else ""
            if piece:
                carried = piece
            elif forward_fill:
                piece = carried
            if piece:
                stacks[c].append(piece)
    return [_norm(" ".join(parts)) for parts in stacks]


def _resolve_columns(
    labels: Sequence[str],
    specs: Sequence[_ColumnSpec],
    *,
    required: Iterable[str],
    where: str,
) -> dict[str, int]:
    """Map spec names onto column indices, or raise saying which failed.

    Two failures, treated differently. A required column that matched
    nothing raises, because the frame would be missing something every
    caller assumes is there. An optional column that matched nothing is
    simply absent from the result — NOT a column of NaN, which reads as
    "the author published blanks" and is a different claim.

    A spec matching MORE than one column always raises, required or
    not. Ambiguity is the dangerous case: picking the first match would
    be a coin flip between two columns whose headers differ by a
    footnote digit.
    """
    required = set(required)
    resolved: dict[str, int] = {}
    for spec in specs:
        hits = [
            i
            for i, label in enumerate(labels)
            if label
            and all(_norm(n) in label for n in spec.needs)
            and not any(_norm(f) in label for f in spec.forbids)
        ]
        if len(hits) == 1:
            resolved[spec.name] = hits[0]
            continue
        if len(hits) > 1:
            raise LongHistoryUnavailable(
                f"{where}: {spec.name!r} matches {len(hits)} columns "
                f"({[labels[i] for i in hits]}). Two columns answering to "
                f"one name means the layout changed; picking one would be "
                f"a guess that reads clean for years."
            )
        if spec.name in required:
            raise LongHistoryUnavailable(
                f"{where}: no column matches {spec.name!r} (needs "
                f"{list(spec.needs)}). The workbook's layout has changed "
                f"and this parser has not. Headers seen: "
                f"{[la for la in labels if la][:24]}"
            )
    return resolved


def _find_row(grid: Grid, predicate: Callable[[list[Any]], bool], *, what: str) -> int:
    for i, row in enumerate(grid):
        try:
            if predicate(row):
                return i
        except Exception:  # a ragged row is not a match
            continue
    raise LongHistoryUnavailable(
        f"could not find {what} anywhere in the sheet. The file's shape "
        f"changed; a parser that guessed at a fixed row number would "
        f"read the wrong rows silently."
    )


def _cell(row: Sequence[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


# -- Shiller --------------------------------------------------------------

SHILLER_SHEET = "Data"

_SHILLER_COLUMNS: tuple[_ColumnSpec, ...] = (
    _ColumnSpec("date", ("date",), ("fraction", "real", "10 year")),
    _ColumnSpec("date_fraction", ("date", "fraction")),
    _ColumnSpec("sp500_price", ("s&p",)),
    _ColumnSpec("dividend", ("dividend",), ("real",)),
    _ColumnSpec("earnings", ("earnings",), ("real", "cape", "ratio", "scaled")),
    _ColumnSpec("cpi", ("cpi",)),
    _ColumnSpec("long_rate", ("gs10",)),
    _ColumnSpec(
        "real_price", ("real", "price"), ("total", "cape", "ratio", "earnings")
    ),
    _ColumnSpec("real_dividend", ("real", "dividend")),
    _ColumnSpec("real_total_return_price", ("real", "total return price")),
    _ColumnSpec(
        "real_earnings", ("real", "earnings"), ("scaled", "cape", "ratio", "bond")
    ),
    _ColumnSpec("real_tr_scaled_earnings", ("scaled", "earnings")),
    _ColumnSpec("cape", ("cape",), ("tr cape", "excess")),
    _ColumnSpec("tr_cape", ("tr cape",)),
    _ColumnSpec("excess_cape_yield", ("excess", "cape", "yield")),
    _ColumnSpec("bond_total_return", ("monthly", "bond", "returns")),
    _ColumnSpec("real_bond_total_return", ("real", "total", "bond", "returns")),
    _ColumnSpec("stock_real_return_10y", ("annualized stock",)),
    _ColumnSpec("bond_real_return_10y", ("annualized bonds",)),
    _ColumnSpec("excess_real_return_10y", ("excess annualized",)),
)

#: Without these there is no series. Everything else in the workbook —
#: CAPE, the total-return variant, the ten-year forward returns — has
#: been added over the years and could reasonably be absent from an
#: archived vintage, so its absence is recorded by the column simply
#: not being in the frame rather than by a raise.
_SHILLER_REQUIRED = (
    "date",
    "sp500_price",
    "dividend",
    "earnings",
    "cpi",
    "long_rate",
)


def parse_shiller(grid: Grid) -> pd.DataFrame:
    """Shiller's monthly sheet as a frame indexed by month start.

    **The price column is a monthly AVERAGE of daily closes, not a
    month-end close.** Shiller says so on his own page and it is the
    single most consequential fact about this dataset. Averaging cannot
    increase the variance of a month-to-month change and it induces
    positive serial correlation: measured on his own real total-return
    series since 1950 the first-order autocorrelation of monthly
    returns is +0.23, against the +0.25 Working (1960) derived for the
    first differences of averages of a pure random walk. A trend or
    momentum rule tested on this series will find persistence that the
    averaging put there. Use it for valuation work and long-horizon
    real returns; do not read a monthly return off it and call it the
    market's monthly return.

    Two smaller ones. Dividends and earnings are interpolated to
    monthly from quarterly (from annual before 1926), so their
    month-to-month variation is arithmetic rather than news. And the
    final row is usually a partial month whose price is a single day's
    close rather than an average — the author says which day in a note
    row at the bottom of the sheet, which `parse_shiller_notes`
    returns rather than discards.

    One repair, and only one. In the months at the end of the file
    where S&P has not reported earnings yet, the nominal earnings cell
    is blank and the REAL earnings cell — a formula deflating it —
    evaluates to a literal 0.0. A zero there is not a quarter in which
    the index earned nothing; it is a missing input with a default
    printed over it, and it is the same shape of mistake as reading
    FRED's holiday `'.'` as a zero interest rate. Left alone it makes
    `real_price / real_earnings` infinite and drags any ten-year
    earnings average down. So an exact zero in a real column is
    returned as NaN WHERE ITS NOMINAL COUNTERPART IS BLANK, and
    nowhere else — that condition is what keeps this from being an
    opinion about the author's numbers.

    Returns a frame with a `date` column of month starts. The month
    sequence is asserted unbroken, which is what catches the
    October/January collapse described in the module docstring.
    """
    header = _find_row(
        grid,
        lambda row: _norm(_cell(row, 0)) == "date",
        what="the Shiller header row (column A reading 'Date')",
    )
    labels = _column_labels(grid, range(max(0, header - 6), header + 1))
    columns = _resolve_columns(
        labels,
        _SHILLER_COLUMNS,
        required=_SHILLER_REQUIRED,
        where="Shiller ie_data.xls",
    )

    date_col = columns["date"]
    fraction_col = columns.get("date_fraction")

    years: list[int] = []
    months: list[int] = []
    values: dict[str, list[float]] = {name: [] for name in columns}
    values.pop("date", None)
    values.pop("date_fraction", None)

    for row in grid[header + 1 :]:
        raw = _cell(row, date_col)
        if _as_float(raw) is None:
            # The sheet ends in prose: the author writes his caveats in
            # the row after the last observation. Stopping at the first
            # non-numeric date is what keeps those out of the series —
            # and `parse_shiller_notes` is where they go instead.
            break
        fraction = _cell(row, fraction_col) if fraction_col is not None else None
        year, month = shiller_month(raw, fraction)
        years.append(year)
        months.append(month)
        for name, index in columns.items():
            if name in ("date", "date_fraction"):
                continue
            values[name].append(coerce_number(_cell(row, index)))

    if not years:
        raise LongHistoryUnavailable(
            "the Shiller sheet had a header row and no numeric dates under "
            "it. That is a layout change or a truncated download, never a "
            "century with nothing in it."
        )

    stamps = pd.to_datetime(
        pd.DataFrame({"year": years, "month": months, "day": 1})
    )
    frame = pd.DataFrame({"date": stamps})
    for name, column in values.items():
        frame[name] = pd.Series(column, dtype="float64")

    frame = _blank_deflated_zeros(frame)
    _assert_unbroken_months(frame["date"], where="Shiller ie_data.xls")
    return frame.reset_index(drop=True)


#: Real columns and the nominal cell each one is deflated from. Only
#: these pairs are eligible for the zero-to-NaN repair in
#: `parse_shiller`, and the nominal side must be blank for it to fire.
_SHILLER_DEFLATED_PAIRS: tuple[tuple[str, str], ...] = (
    ("earnings", "real_earnings"),
    ("earnings", "real_tr_scaled_earnings"),
    ("dividend", "real_dividend"),
)


def _blank_deflated_zeros(frame: pd.DataFrame) -> pd.DataFrame:
    """NaN a real column's zero where the nominal cell it deflates is blank."""
    for nominal, real in _SHILLER_DEFLATED_PAIRS:
        if nominal not in frame.columns or real not in frame.columns:
            continue
        artefact = frame[nominal].isna() & (frame[real] == 0.0)
        if artefact.any():
            frame.loc[artefact, real] = float("nan")
    return frame


def parse_shiller_notes(grid: Grid) -> pd.DataFrame:
    """The prose rows under the data, which carry live caveats.

    Not decoration. The vintage read on 2026-08-02 ends with "July
    price is July 7th close", "Oct '25/June/July CPI estimated" and
    "July GS10 is July 1st value" — three statements about the last
    rows of the series that exist nowhere else, and that a parser
    stopping at the first non-numeric date would throw away. Returned
    as `row`, `column`, `text`.
    """
    header = _find_row(
        grid,
        lambda row: _norm(_cell(row, 0)) == "date",
        what="the Shiller header row (column A reading 'Date')",
    )
    started = False
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(grid[header + 1 :], start=header + 1):
        numeric = _as_float(_cell(row, 0)) is not None
        if numeric:
            started = True
            continue
        if not started:
            continue
        for c, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip():
                rows.append({"row": i, "column": c, "text": cell.strip()})
    return pd.DataFrame(rows, columns=["row", "column", "text"])


def _assert_unbroken_months(dates: pd.Series, *, where: str) -> None:
    """Every month present exactly once, first to last.

    This is the check that actually catches the YYYY.MM trap. A parser
    that read 1871.10 as January produces two Januaries and no October,
    which leaves the row count unchanged, the dtypes correct and every
    value in the right column — and shows up here as a duplicate and a
    gap.
    """
    index = pd.DatetimeIndex(dates)
    if len(index) == 0:
        raise LongHistoryUnavailable(f"{where}: no dated rows")
    expected = pd.date_range(index[0], index[-1], freq="MS")
    if index.equals(expected):
        return

    duplicated = sorted({str(d.date()) for d in index[index.duplicated()]})
    missing = sorted({str(d.date()) for d in expected.difference(index)})
    raise LongHistoryUnavailable(
        f"{where}: the parsed months are not a complete monthly sequence "
        f"from {index[0].date()} to {index[-1].date()}. "
        f"{len(duplicated)} duplicated {duplicated[:8]}, "
        f"{len(missing)} missing {missing[:8]}. The counts are printed "
        f"beside the samples because the samples are truncated and a "
        f"long list hides the one month that would name the cause. The "
        f"classic cause is reading the YYYY.MM date column as a decimal "
        f"year, which folds every October (stored as the float YYYY.1) "
        f"onto January and leaves exactly this pattern."
    )


# -- Damodaran: annual returns -------------------------------------------

DAMODARAN_RETURNS_SHEET = "Returns by year"

#: Read against the combined "block | column" label. The pipe is load
#: bearing in three places. `| year` anchors the year column, because
#: "10-year" contains "year" in four other headers. And `1928 in |`
#: versus `1928 in real terms |` is the only thing separating the
#: nominal wealth block from the real one — the nominal block's name is
#: a strict prefix of the real block's, so a needle without the
#: separator matches both and the resolver would report an ambiguity
#: rather than, much worse, silently take the first.
_RETURNS = "annual returns on investments in |"
_WEALTH = "1928 in |"
_PREMIUM = "annual risk premium |"
_REAL_RETURNS = "annual real returns |"
_REAL_WEALTH = "1928 in real terms |"
_REAL_PREMIUM = "real risk premium |"

#: The seven asset labels, in the order the sheet writes them. Named
#: once because they repeat across five blocks, and a typo in the
#: fourth copy would silently resolve to a different asset's column.
_DAM_ASSETS: tuple[tuple[str, str], ...] = (
    ("sp500", "s&p 500"),
    ("smallcap", "small cap"),
    ("tbill", "t.bill"),
    ("tbond", "t.bond"),
    ("baa", "baa"),
    ("real_estate", "real estate"),
    ("gold", "gold"),
)


def _asset_specs(block: str, suffix: str) -> tuple[_ColumnSpec, ...]:
    return tuple(
        _ColumnSpec(f"{name}_{suffix}", (block, needle))
        for name, needle in _DAM_ASSETS
    )


_DAMODARAN_RETURN_COLUMNS: tuple[_ColumnSpec, ...] = (
    _ColumnSpec("year", ("| year",)),
    *_asset_specs(_RETURNS, "return"),
    *_asset_specs(_WEALTH, "value"),
    *_asset_specs(_REAL_RETURNS, "real_return"),
    *_asset_specs(_REAL_WEALTH, "real_value"),
    # Premiums, nominal.
    _ColumnSpec("stocks_minus_bills", (_PREMIUM, "stocks - bills")),
    _ColumnSpec("stocks_minus_bonds", (_PREMIUM, "stocks - bonds")),
    _ColumnSpec("smallcap_premium", (_PREMIUM, "small cap premium")),
    _ColumnSpec("stocks_minus_baa", (_PREMIUM, "baa")),
    # Premiums, real.
    _ColumnSpec("stocks_minus_bills_real", (_REAL_PREMIUM, "t.bills")),
    _ColumnSpec("stocks_minus_bonds_real", (_REAL_PREMIUM, "t.bonds")),
    # `inflation` sits in the real-returns block beside the seven
    # assets and is not one of them.
    _ColumnSpec("inflation", (_REAL_RETURNS, "inflation")),
)

_DAMODARAN_RETURN_REQUIRED = (
    "year",
    "sp500_return",
    "tbill_return",
    "tbond_return",
    "inflation",
)

#: Deliberately not parsed. `Historical ERP` in the risk-premium block
#: is a live spreadsheet formula wired to the "Enter your starting
#: year" input cell near the top of the sheet, so its cached value is
#: whatever the last person to open the file happened to type — not a
#: published series. It is blank in the vintage read here, and a blank
#: column would have been read as "he stopped publishing it".
DAMODARAN_INTERACTIVE_COLUMNS = ("Historical ERP",)


def parse_damodaran_annual(grid: Grid) -> pd.DataFrame:
    """The 1928-onward annual return table as a frame keyed by year.

    Nominal and real returns for the S&P 500 with dividends, the
    bottom-decile small-cap portfolio, 3-month bills, the 10-year
    Treasury, Baa corporates, residential property and gold; the value
    of $100 compounded in each since 1928; and the realised risk
    premiums.

    Note what the T-bond column is: not an observed total-return index
    but Damodaran's own coupon-and-price arithmetic applied to the
    constant-maturity yield. It behaves like a bond return and it is a
    construction, so a bond result computed from it is partly a test of
    the construction.
    """
    header = _find_row(
        grid,
        lambda row: _norm(_cell(row, 0)) == "year",
        what="the Damodaran header row (column A reading 'Year')",
    )
    labels = _combined_labels(grid, banner=header - 1, header=header)
    columns = _resolve_columns(
        labels,
        _DAMODARAN_RETURN_COLUMNS,
        required=_DAMODARAN_RETURN_REQUIRED,
        where="Damodaran histretSP.xls",
    )
    return _annual_rows(grid, header, columns, where="Damodaran histretSP.xls")


# -- Damodaran: implied ERP, annual --------------------------------------

DAMODARAN_IMPLIED_ERP_SHEET = "Historical Impl Premiums"

_DAMODARAN_ERP_COLUMNS: tuple[_ColumnSpec, ...] = (
    _ColumnSpec("year", ("| year",)),
    _ColumnSpec("earnings_yield", ("earnings yield",)),
    _ColumnSpec("dividend_yield", ("dividend yield",)),
    _ColumnSpec("sp500", ("s&p 500",)),
    _ColumnSpec("earnings", ("earnings*",)),
    _ColumnSpec("dividends", ("dividends*",)),
    _ColumnSpec("dividends_and_buybacks", ("dividends + buybacks",)),
    _ColumnSpec("change_in_earnings", ("change in earnings",)),
    _ColumnSpec("change_in_dividends", ("change in dividends",)),
    _ColumnSpec("tbill_rate", ("t.bill rate",)),
    _ColumnSpec("tbond_rate", ("t.bond rate",)),
    _ColumnSpec("bond_minus_bill", ("bond-bill",)),
    _ColumnSpec("smoothed_growth", ("smoothed growth",)),
    _ColumnSpec("implied_premium_ddm", ("implied premium (ddm)",)),
    _ColumnSpec("analyst_growth", ("analyst growth",)),
    _ColumnSpec("implied_erp_fcfe", ("implied erp (fcfe)",)),
    _ColumnSpec("implied_erp_adjusted_riskfree", ("implied erp with risk adjusted",)),
    _ColumnSpec("implied_premium_sustainable_payout", ("sustainable payout",)),
    _ColumnSpec("erp_over_riskfree", ("erp/riskfree",)),
)


def parse_damodaran_implied_erp_annual(grid: Grid) -> pd.DataFrame:
    """Year-end implied equity risk premiums, 1960 onward.

    A discount rate rather than a return: the premium that would
    justify the index level given its cash yield, an expected growth
    rate and the long bond. Two consequences worth stating where
    somebody will read them. It is not a forecast of what equities
    returned, and comparing it to `damodaran_annual`'s realised premium
    is comparing two different quantities that happen to share a name.
    And it starts in 1960, so it is not century-scale and cannot say
    anything about 1929 or 1974.
    """
    header = _find_row(
        grid,
        lambda row: _norm(_cell(row, 0)) == "year",
        what="the implied-ERP header row (column A reading 'Year')",
    )
    labels = _combined_labels(grid, banner=None, header=header)
    columns = _resolve_columns(
        labels,
        _DAMODARAN_ERP_COLUMNS,
        required=("year", "implied_premium_ddm"),
        where="Damodaran histimpl.xls",
    )
    return _annual_rows(grid, header, columns, where="Damodaran histimpl.xls")


# -- Damodaran: implied ERP, monthly -------------------------------------

DAMODARAN_ERP_MONTHLY_SHEET = "Historical ERP"

_DAMODARAN_ERP_MONTHLY_COLUMNS: tuple[_ColumnSpec, ...] = (
    _ColumnSpec("date", ("start of month",)),
    _ColumnSpec("sp500", ("s&p 500",)),
    _ColumnSpec("tbond_rate", ("t.bond rate",)),
    _ColumnSpec("riskfree_rate", ("$ riskfree rate",)),
    _ColumnSpec("ten_year_average_cashflow", ("ten-year average cf",)),
    _ColumnSpec("trailing_cashflow", ("cf (trailing 12 month)",)),
    _ColumnSpec("normalized_cashflow", ("normalized cf",)),
    _ColumnSpec("expected_growth", ("expected growth rate",)),
    _ColumnSpec(
        "erp_sustainable_payout", ("erp (t12 m with sustainable payout)",)
    ),
    _ColumnSpec("erp_trailing", ("erp (t12m)",), ("adj riskfree",)),
    _ColumnSpec("erp_trailing_adjusted_riskfree", ("erp (t12m) with adj riskfree",)),
    _ColumnSpec("erp_smoothed", ("erp (smoothed)",)),
    _ColumnSpec("erp_normalized", ("erp (normalized)",)),
    _ColumnSpec("erp_net_cash_yield", ("net cash yield",)),
    _ColumnSpec("erp_covid_adjusted", ("covid adjusted",)),
    _ColumnSpec("expected_return", ("expected return",)),
    _ColumnSpec("notes", ("notes",), text=True),
)


def parse_damodaran_erp_monthly(grid: Grid) -> pd.DataFrame:
    """Monthly implied ERP, September 2008 onward.

    The shortest series in this module by a wide margin — it is here
    because the brief asked for the implied premium, not because
    eighteen years tests anything the rest of our sample does not
    already cover.

    It is also the only hand-typed one. Roughly one numeric cell in a
    hundred arrives as a string, including one month where every field
    does, and `coerce_number` repairs what it can. The repairs are
    counted and warned about: a number that jumps means the author's
    formatting changed and somebody should look at the sheet rather
    than at the series.
    """
    header = _find_row(
        grid,
        lambda row: _norm(_cell(row, 0)) == "start of month",
        what="the monthly-ERP header row (column A reading 'Start of month')",
    )
    labels = _combined_labels(grid, banner=None, header=header)
    columns = _resolve_columns(
        labels,
        _DAMODARAN_ERP_MONTHLY_COLUMNS,
        required=("date", "erp_trailing"),
        where="Damodaran ERPbymonth.xlsx",
    )
    text_columns = {s.name for s in _DAMODARAN_ERP_MONTHLY_COLUMNS if s.text}

    date_col = columns["date"]
    stamps: list[pd.Timestamp] = []
    values: dict[str, list[Any]] = {n: [] for n in columns if n != "date"}
    repaired: dict[str, int] = {}

    for row in grid[header + 1 :]:
        stamp = _month_start(_cell(row, date_col))
        if stamp is None:
            break
        stamps.append(stamp)
        for name, index in columns.items():
            if name == "date":
                continue
            cell = _cell(row, index)
            if name in text_columns:
                values[name].append(
                    cell.strip() if isinstance(cell, str) and cell.strip() else None
                )
                continue
            number = coerce_number(cell)
            if isinstance(cell, str) and cell.strip() and not math.isnan(number):
                repaired[name] = repaired.get(name, 0) + 1
            values[name].append(number)

    if not stamps:
        raise LongHistoryUnavailable(
            "the monthly ERP sheet had a header row and no dated rows "
            "under it."
        )

    frame = pd.DataFrame({"date": pd.DatetimeIndex(stamps)})
    for name, column in values.items():
        if name in text_columns:
            frame[name] = pd.Series(column, dtype="object")
        else:
            frame[name] = pd.Series(column, dtype="float64")

    if repaired:
        total = sum(repaired.values())
        warnings.warn(
            f"Damodaran ERPbymonth: repaired {total} number(s) typed as "
            f"text ({repaired}). That is normal for this sheet — a "
            f"percent sign, a European decimal comma, a footnote "
            f"asterisk — and a sharp rise in the count means the "
            f"formatting changed rather than the market.",
            stacklevel=2,
        )

    _assert_unbroken_months(frame["date"], where="Damodaran ERPbymonth.xlsx")
    return frame.reset_index(drop=True)


# -- shared row readers ---------------------------------------------------


def _combined_labels(grid: Grid, *, banner: int | None, header: int) -> list[str]:
    """`"<block> | <column>"` per column, with the block forward-filled.

    Where there is no banner row the block is empty and the label is
    `"| earnings yield"`. Uniform on purpose: the `| year` anchor used
    by both Damodaran parsers then means the same thing in a sheet with
    blocks and a sheet without.
    """
    heads = _column_labels(grid, [header])
    if banner is None or banner < 0:
        return [f"| {h}" for h in heads]
    blocks = _column_labels(grid, [banner], forward_fill=True)
    width = max(len(heads), len(blocks))
    heads += [""] * (width - len(heads))
    blocks += [""] * (width - len(blocks))
    return [f"{b} | {h}" for b, h in zip(blocks, heads)]


def _annual_rows(
    grid: Grid, header: int, columns: Mapping[str, int], *, where: str
) -> pd.DataFrame:
    """Rows under `header` while column `year` still holds a year.

    Damodaran ends both annual sheets with summary blocks — arithmetic
    and geometric averages over several windows — whose first column is
    a period label like `1928-2025`. Stopping at the first row whose
    year cell is not a plausible year is what keeps a 98-year average
    out of the series as though it were a 99th year.
    """
    year_col = columns["year"]
    years: list[int] = []
    values: dict[str, list[float]] = {n: [] for n in columns if n != "year"}

    for row in grid[header + 1 :]:
        raw = _as_float(_cell(row, year_col))
        if raw is None or not 1500 <= raw <= 2200 or raw != int(raw):
            break
        years.append(int(raw))
        for name, index in columns.items():
            if name == "year":
                continue
            values[name].append(coerce_number(_cell(row, index)))

    if not years:
        raise LongHistoryUnavailable(
            f"{where}: a header row with no year rows under it."
        )
    if len(set(years)) != len(years) or years != sorted(years):
        raise LongHistoryUnavailable(
            f"{where}: the year column is not strictly increasing "
            f"({years[:6]}...). A repeated or out-of-order year means the "
            f"summary block was read as data."
        )

    frame = pd.DataFrame({"year": pd.Series(years, dtype="int64")})
    for name, column in values.items():
        frame[name] = pd.Series(column, dtype="float64")
    return frame


def _month_start(value: Any) -> pd.Timestamp | None:
    """A cell holding a month, normalised to the first of that month."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value.year, value.month, 1)
    if isinstance(value, str) and value.strip():
        stamp = pd.to_datetime(value.strip(), errors="coerce")
        if pd.isna(stamp):
            return None
        return pd.Timestamp(stamp.year, stamp.month, 1)
    return None


# -- derived views --------------------------------------------------------


def annual_from_shiller(monthly: pd.DataFrame, *, month: int = 12) -> pd.DataFrame:
    """Year-over-year total returns from Shiller's monthly series.

    Read the caveat before the numbers: Shiller's December price is an
    average of December's daily closes, so a December-to-December
    return here is roughly a year ending in mid-December, not a
    calendar year. It is close enough to compare against an annual
    compilation and it is not the same quantity, which is exactly what
    `reconcile_annual_equity_return` is for.

    Returns `year`, `real_total_return`, `nominal_total_return`,
    `inflation`.
    """
    needed = {"date", "real_total_return_price", "cpi"}
    missing = sorted(needed - set(monthly.columns))
    if missing:
        raise ValueError(f"annual_from_shiller needs {missing}")

    dated = monthly.loc[pd.DatetimeIndex(monthly["date"]).month == month]
    dated = dated.sort_values("date")
    if len(dated) < 2:
        raise ValueError(
            f"only {len(dated)} observation(s) for month {month}; an annual "
            f"return needs two consecutive ones"
        )

    years = pd.DatetimeIndex(dated["date"]).year.to_numpy()
    real = dated["real_total_return_price"].to_numpy(dtype="float64")
    cpi = dated["cpi"].to_numpy(dtype="float64")
    nominal = real * cpi

    out = pd.DataFrame(
        {
            "year": years[1:],
            "real_total_return": real[1:] / real[:-1] - 1.0,
            "nominal_total_return": nominal[1:] / nominal[:-1] - 1.0,
            "inflation": cpi[1:] / cpi[:-1] - 1.0,
        }
    )
    # A year whose December is missing leaves a two-year gap that would
    # otherwise be reported as a one-year return of about 21%.
    contiguous = out["year"].to_numpy() - years[:-1] == 1
    return out.loc[contiguous].reset_index(drop=True)


def reconcile_annual_equity_return(
    shiller_monthly: pd.DataFrame,
    damodaran: pd.DataFrame,
    *,
    tolerance: float = 0.05,
) -> pd.DataFrame:
    """Where two independent compilations of the same market disagree.

    Every internal check in this repository compares data against
    itself, and a consistently wrong series passes all of them. Shiller
    and Damodaran built their US equity histories separately and both
    cover 1928-2025, so a bar-for-bar diff is the cheapest real
    evidence available for either.

    What agreement looks like, measured on the vintages read on
    2026-08-02: correlation 0.989 across 98 years, median absolute
    annual difference 1.7 percentage points, worst year 8.2 points
    (1991), and geometric means of 10.18% against 10.02% — sixteen
    basis points apart over 98 years. The year-by-year gap is not an
    error in either: Shiller's December is an average of December's
    daily closes and Damodaran's year ends on the 31st.

    So `tolerance` is a question about what you are looking for. The
    default of 0.05 surfaces the handful of genuinely divergent years.
    Set it to 0.20 and a non-empty result means a parser is reading the
    wrong column, which is what this function is really insurance
    against.

    Returns `year`, `shiller`, `damodaran`, `diff`, sorted by year.
    Raises when the two share no years — an empty frame there would be
    indistinguishable from perfect agreement.
    """
    if tolerance < 0:
        raise ValueError(f"tolerance must be non-negative, got {tolerance}")
    if "sp500_return" not in damodaran.columns:
        raise ValueError("damodaran frame has no 'sp500_return' column")

    left = annual_from_shiller(shiller_monthly).set_index("year")
    right = damodaran.set_index("year")
    shared = left.index.intersection(right.index)
    if len(shared) == 0:
        raise ValueError(
            "these two frames share no years "
            f"(shiller {left.index.min()}-{left.index.max()}, "
            f"damodaran {right.index.min()}-{right.index.max()}). Returning "
            "an empty result would be indistinguishable from two sources "
            "agreeing on every year."
        )

    a = left.loc[shared, "nominal_total_return"].astype("float64")
    b = right.loc[shared, "sp500_return"].astype("float64")
    out = pd.DataFrame(
        {
            "year": shared.astype("int64"),
            "shiller": a.to_numpy(),
            "damodaran": b.to_numpy(),
        }
    )
    out["diff"] = out["shiller"] - out["damodaran"]
    flagged = out.loc[out["diff"].abs() > tolerance]
    return flagged.sort_values("year").reset_index(drop=True)


# -- the library ----------------------------------------------------------


class LongHistory:
    """Fetches, caches and parses the long-horizon files.

    The session, the sleep and the clock are constructor arguments so a
    test can stand in for all three; nothing here reaches the network
    at construction time, and a warm cache is readable with no network
    at all — which is the state a reviewer rerunning a report is in.
    """

    def __init__(
        self,
        cache: "ParquetCache | None" = None,
        *,
        session: requests.Session | None = None,
        timeout: float = 120.0,
        sleep: Callable[[float], None] = time.sleep,
        clock: Callable[[], datetime] | None = None,
    ) -> None:
        self._cache = cache
        self._session = session or requests.Session()
        self._timeout = timeout
        self._sleep = sleep
        self._clock = clock or (lambda: datetime.now(timezone.utc))
        self._bytes: dict[str, bytes] = {}
        self._last_request: float = float("-inf")

    # -- raw files ------------------------------------------------------

    def raw(self, key: str, *, refresh: bool = False) -> bytes:
        """The workbook itself, from cache or from the author's server.

        Cached on the URL alone with no expiry. A file its author
        updates once each January, re-downloaded because a timer fired,
        is a request that returns no information and costs goodwill;
        `refresh=True` is how a new vintage is asked for, and it is
        meant to be a decision somebody made rather than a default
        nobody saw.
        """
        dataset = self._dataset(key)
        if not refresh and key in self._bytes:
            return self._bytes[key]

        def load() -> pd.DataFrame:
            payload = self._download(dataset.url)
            return pd.DataFrame(
                {
                    "url": [dataset.url],
                    "sha256": [hashlib.sha256(payload).hexdigest()],
                    "bytes": [len(payload)],
                    "retrieved": [self._clock().replace(tzinfo=None)],
                    "content": [payload],
                }
            )

        if self._cache is None:
            frame = load()
        else:
            cache_key = self._cache.key(
                "longhistory", "workbook", dataset=key, url=dataset.url
            )
            frame = self._cache.get_or_load(
                cache_key,
                load,
                stamped=self._clock(),
                # No `now`, so nothing here ever expires on a timer.
                now=None,
                refresh=refresh,
            )

        payload = bytes(frame["content"].iloc[0])
        self._bytes[key] = payload
        return payload

    def digest(self, key: str, *, refresh: bool = False) -> str:
        """sha256 of the cached workbook — the vintage id.

        Worth recording next to any result computed from Damodaran's
        files, since he revises method between years and the numbers
        for 1928 are not promised to be stable across vintages.
        """
        return hashlib.sha256(self.raw(key, refresh=refresh)).hexdigest()

    # -- parsed frames --------------------------------------------------

    def shiller(self, *, refresh: bool = False) -> pd.DataFrame:
        grid = grid_from_xls(self.raw("shiller", refresh=refresh), SHILLER_SHEET)
        return parse_shiller(grid)

    def shiller_notes(self, *, refresh: bool = False) -> pd.DataFrame:
        grid = grid_from_xls(self.raw("shiller", refresh=refresh), SHILLER_SHEET)
        return parse_shiller_notes(grid)

    def damodaran_returns(self, *, refresh: bool = False) -> pd.DataFrame:
        grid = grid_from_xls(
            self.raw("damodaran_returns", refresh=refresh),
            DAMODARAN_RETURNS_SHEET,
        )
        return parse_damodaran_annual(grid)

    def damodaran_implied_erp(self, *, refresh: bool = False) -> pd.DataFrame:
        grid = grid_from_xls(
            self.raw("damodaran_implied_erp", refresh=refresh),
            DAMODARAN_IMPLIED_ERP_SHEET,
        )
        return parse_damodaran_implied_erp_annual(grid)

    def damodaran_implied_erp_monthly(
        self, *, refresh: bool = False
    ) -> pd.DataFrame:
        grid = grid_from_xlsx(
            self.raw("damodaran_erp_monthly", refresh=refresh),
            DAMODARAN_ERP_MONTHLY_SHEET,
        )
        return parse_damodaran_erp_monthly(grid)

    # -- FRED -----------------------------------------------------------

    def fred_macro(
        self,
        series: Iterable[str] = tuple(FRED_CENTURY_SERIES),
        *,
        start: date = date(1850, 1, 1),
        end: date | None = None,
        refresh: bool = False,
    ) -> pd.DataFrame:
        """Long US macro series in LONG form: `series`, `date`, `value`.

        Long rather than wide, and that is a decision about honesty
        rather than about taste. These series begin in eight different
        decades; a wide frame fills the years before each one starts
        with NaN, and a NaN that means "not published until 1948" is
        then indistinguishable from a NaN that means "published as
        missing". In long form a month that does not exist simply has
        no row. Pivot with `.pivot(index="date", columns="series",
        values="value")` once you have decided what an absence means
        for the thing you are computing.

        USREC needs saying twice: NBER assigns recession dates
        retrospectively, often more than a year after the fact, so the
        value for a month was not knowable in that month. It is fine
        for describing history and must never be an input to a rule
        that trades.
        """
        wanted = tuple(dict.fromkeys(str(s).strip().upper() for s in series))
        if not wanted:
            raise ValueError("no FRED series requested")
        through = end or self._clock().date()

        def load() -> pd.DataFrame:
            frames: list[pd.DataFrame] = []
            for i, name in enumerate(wanted):
                if i:
                    # FRED tolerates far more than this; the pause is
                    # here because nine series in a tight loop is the
                    # shape of a request pattern that gets an address
                    # blocked, and the whole pull is cached anyway.
                    self._sleep(1.0)
                try:
                    obs = fetch_observations(name, start, through)
                except TbillUnavailable as exc:
                    raise LongHistoryUnavailable(
                        f"FRED did not serve {name}: {exc}. That is our "
                        f"outage and not a statement about the American "
                        f"economy — do not let a macro panel come back "
                        f"short on the strength of it."
                    ) from exc
                frames.append(
                    pd.DataFrame(
                        {
                            "series": name,
                            "date": pd.DatetimeIndex(obs["date"]),
                            "value": obs["value"].astype("float64"),
                        }
                    )
                )
            out = pd.concat(frames, ignore_index=True)
            return out.sort_values(["series", "date"]).reset_index(drop=True)

        if self._cache is None:
            return load()

        # `end` is in the key, so a new day is a new entry and a
        # same-day rerun is free. That is why nothing here needs a TTL:
        # the key already carries the only thing that would have
        # expired.
        cache_key = self._cache.key(
            "longhistory",
            "fred_macro",
            series=list(wanted),
            start=start,
            end=through,
        )
        return self._cache.get_or_load(
            cache_key,
            load,
            stamped=self._clock(),
            now=None,
            refresh=refresh,
        )

    # -- description ----------------------------------------------------

    @staticmethod
    def catalogue() -> pd.DataFrame:
        return catalogue()

    @staticmethod
    def describe(key: str) -> str:
        """A dataset's honest paragraph, for a report or a docstring."""
        dataset = DATASETS.get(key)
        if dataset is None:
            raise KeyError(f"unknown dataset {key!r}; have {sorted(DATASETS)}")
        lines = [
            f"{dataset.name}",
            f"  source      {dataset.author}",
            f"  url         {dataset.url}",
            f"  frequency   {dataset.frequency}",
            f"  starts      {dataset.starts}",
            f"  through     {dataset.observed_through} (vintage read 2026-08-02)",
            f"  updated     {dataset.cadence}",
            f"  terms       {dataset.redistribution}",
            f"  survivorship {dataset.survivorship}",
            "  can:",
            *(f"    - {c}" for c in dataset.can),
            "  cannot:",
            *(f"    - {c}" for c in dataset.cannot),
        ]
        return "\n".join(lines)

    # -- transport ------------------------------------------------------

    @staticmethod
    def _dataset(key: str) -> LongDataset:
        dataset = DATASETS.get(key)
        if dataset is None:
            raise KeyError(f"unknown dataset {key!r}; have {sorted(DATASETS)}")
        return dataset

    def _download(self, url: str) -> bytes:
        last_reason = "no attempt was made"

        for attempt in range(MAX_ATTEMPTS):
            if attempt:
                self._sleep(
                    min(
                        BACKOFF_BASE_SECONDS * 2 ** (attempt - 1),
                        BACKOFF_CAP_SECONDS,
                    )
                )
            else:
                self._pace()
            try:
                response = self._session.get(
                    url,
                    headers={"User-Agent": USER_AGENT},
                    timeout=self._timeout,
                )
            except requests.RequestException as exc:
                last_reason = f"{type(exc).__name__}: {exc}"
                continue

            status = response.status_code
            if status == 200:
                payload = response.content
                _require_workbook(payload, url)
                return payload
            if status == 404:
                raise LongHistoryUnavailable(
                    f"{url} is gone (404). These are personal academic "
                    f"pages and they get reorganised; find the new link on "
                    f"the author's data page and update the constant rather "
                    f"than falling back to a mirror, which for Shiller means "
                    f"silently losing three years of history."
                )
            if status in RETRY_STATUSES:
                last_reason = f"HTTP {status}"
                continue
            raise LongHistoryUnavailable(f"{url} returned HTTP {status}")

        raise LongHistoryUnavailable(
            f"{url} unreachable after {MAX_ATTEMPTS} attempts. Last: "
            f"{last_reason}. No rows were returned and none were claimed — "
            f"this is an outage, not a century with no data in it."
        )

    def _pace(self) -> None:
        """Keep our own burst under what a personal web server deserves.

        `monotonic` rather than the injected clock: that one exists so
        a cache stamp can be made deterministic in a test, and a frozen
        clock must not quietly become a decision about how fast we are
        allowed to talk to somebody else's machine.
        """
        waited = time.monotonic() - self._last_request
        if 0.0 <= waited < MIN_REQUEST_INTERVAL_SECONDS:
            self._sleep(MIN_REQUEST_INTERVAL_SECONDS - waited)
        self._last_request = time.monotonic()


def _require_workbook(payload: bytes, url: str) -> None:
    """Refuse anything that is not an Excel file, before parsing it.

    A personal web host's characteristic failure is HTTP 200 carrying
    an error page, a cookie wall or a redirect notice. Letting that
    through means the traceback names xlrd, and whoever reads it goes
    looking at the parser rather than at the twelve hundred bytes of
    HTML we were served.
    """
    if not payload:
        raise LongHistoryUnavailable(f"{url} returned an empty body")
    if payload.startswith(_OLE2_MAGIC) or payload.startswith(_ZIP_MAGIC):
        return
    head = payload[:200].decode("utf-8", "replace").strip()
    raise LongHistoryUnavailable(
        f"{url} answered 200 with something that is not an Excel workbook "
        f"({len(payload)} bytes beginning {head[:120]!r}). A 200 carrying "
        f"an error page is how these hosts report trouble, and reading it "
        f"as a bad file would send somebody to debug the parser."
    )
